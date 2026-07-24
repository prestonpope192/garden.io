#!/usr/bin/env python3
"""Generate draft catalog plant JSON records from the starter workbook.

This is the deterministic first pass for batch catalog population. It does not
invent horticultural facts. It preserves workbook provenance, fills every
top-level section required by the import contract, and marks unknown detail as
draft/low-confidence so source-enrichment can happen in a later pass.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path("/Users/preston/Downloads/garden_io_starter_import_march_2026.xlsx")
DEFAULT_OUTPUT_DIR = Path("data/catalog/plant-records/starter-batch")
SCHEMA_VERSION = "2026-06-plant-profile-v1"


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "plant"


def display_name_for(row: dict[str, Any], slug: str) -> str:
    raw_name = clean_text(row.get("plant_name")) or slug.replace("-", " ").title()
    name = re.sub(r"\s+replacement$", "", raw_name, flags=re.IGNORECASE)
    if "—" in name:
        generic, cultivar = [part.strip() for part in name.split("—", 1)]
        if slugify(generic) == slug:
            return generic
        return f"{generic} ({cultivar})"
    return name


def infer_plant_type(name: str, slug: str) -> str:
    text = f"{name} {slug}".lower()
    if any(word in text for word in ["apple", "peach", "nectarine", "pecan", "yuzu", "satsuma", "lemon", "lime", "orange", "loquat", "guava", "pawpaw", "jujube", "pear"]):
        return "tree"
    if any(word in text for word in ["blueberry", "gooseberry", "currant", "raspberry", "blackberry", "lingonberry", "salal"]):
        return "shrub"
    if any(word in text for word in ["onion", "chive", "dill", "borage", "nasturtium", "comfrey", "bee balm"]):
        return "herb"
    if "strawberry" in text:
        return "groundcover"
    return "forb"


def read_sheet(workbook: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(workbook, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell).strip() for cell in rows[0] if cell is not None]
    records: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        values = raw_row[: len(headers)]
        if not any(value not in (None, "") for value in values):
            continue
        records.append(dict(zip(headers, values)))
    return records


def workbook_candidates(workbook: Path, prefer_explicit_slugs: bool) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet_name in ["plants", "wishlist"]:
        for row in read_sheet(workbook, sheet_name):
            name = clean_text(row.get("plant_name"))
            if not name:
                continue
            explicit_slug = clean_text(row.get("catalog_slug"))
            if prefer_explicit_slugs and not explicit_slug:
                continue
            slug = slugify(explicit_slug or name)
            rows.append({**row, "_sheet_name": sheet_name, "_slug": slug})

    seen: set[str] = set()
    unique_rows: list[dict[str, Any]] = []
    for row in rows:
        slug = row["_slug"]
        if slug in seen:
            continue
        seen.add(slug)
        unique_rows.append(row)
    return unique_rows


def source_note(row: dict[str, Any]) -> str:
    parts = [
        clean_text(row.get("source")),
        f"Workbook sheet: {row.get('_sheet_name')}.",
    ]
    if clean_text(row.get("zone_name")) or clean_text(row.get("candidate_zone")):
        parts.append(f"Candidate/actual zone: {clean_text(row.get('zone_name')) or clean_text(row.get('candidate_zone'))}.")
    if clean_text(row.get("bed_name")) or clean_text(row.get("candidate_bed")):
        parts.append(f"Candidate/actual bed: {clean_text(row.get('bed_name')) or clean_text(row.get('candidate_bed'))}.")
    if clean_text(row.get("notes")):
        parts.append(f"Workbook notes: {clean_text(row.get('notes'))}")
    return " ".join(part for part in parts if part)


def draft_record(row: dict[str, Any]) -> dict[str, Any]:
    slug = row["_slug"]
    display_name = display_name_for(row, slug)
    source_ref = f"starter-workbook-{slug}"
    plant_type = infer_plant_type(display_name, slug)
    now = datetime.now(timezone.utc).isoformat()

    return {
        "schema_version": SCHEMA_VERSION,
        "slug": slug,
        "taxonomy": {
            "kingdom_name": "Plantae",
            "family_name": None,
            "genus_name": slug.replace("-", "_"),
            "species_name": None,
            "subspecies_name": None,
            "variety_name": None,
            "botanical_name_full": display_name,
            "taxon_rank": "unknown",
            "native_range": None,
            "origin_type": "unknown",
        },
        "names": [
            {
                "name": display_name,
                "name_type": "common",
                "locale": "en",
                "is_primary": True,
            }
        ],
        "profile": {
            "display_name": display_name,
            "plant_type_code": plant_type,
            "lifecycle_type": "unknown",
            "confidence_score": 0.2,
            "evidence_count": 1,
            "source_count": 1,
            "source_last_reviewed_at": now,
            "ai_generated_summary": False,
            "human_verified": False,
            "conflict_flag": False,
            "region_specific_conflict_notes": None,
            "is_ai_generated": False,
            "generation_status": "community_generated",
            "is_published": False,
            "review_status": "draft",
        },
        "aesthetic_styles": [],
        "uses": [],
        "narratives": {
            "locale": "en",
            "short_description": f"{display_name} imported as a draft catalog plant from the March 2026 Garden.io starter workbook.",
            "why_plant_it": None,
            "pros_summary": None,
            "cons_summary": None,
            "primary_use_cases": None,
            "notes_for_homestead": source_note(row),
            "notes_for_small_garden": None,
            "notes_for_container_growing": None,
            "editorial_summary": None,
        },
        "ornamental": {
            "evergreen_deciduous": "unknown",
            "ornamental_season_interest": [],
            "visual_texture": None,
            "foliage_color": None,
            "evergreen_foliage": None,
            "winter_interest": None,
        },
        "climate": {
            "usda_hardiness_min": None,
            "usda_hardiness_max": None,
            "ahs_heat_zone_min": None,
            "ahs_heat_zone_max": None,
            "cold_tolerance_absolute_f": None,
            "cold_tolerance_established_f": None,
            "heat_tolerance_f": None,
            "humidity_tolerance_code": "unknown",
            "drought_tolerance_code": "unknown",
            "flood_tolerance_code": "unknown",
            "wind_tolerance_code": "unknown",
            "salt_tolerance_code": "unknown",
            "chill_hours_min": None,
            "chill_hours_max": None,
            "frost_tender": None,
            "reemergence_after_freeze_behavior": None,
            "sun_min_hours": None,
            "sun_max_hours": None,
            "preferred_light": None,
            "shade_tolerance_score": None,
            "afternoon_sun_tolerance_score": None,
            "reflected_heat_tolerance_score": None,
        },
        "growth": {
            "mature_height_min_in": None,
            "mature_height_max_in": None,
            "mature_width_min_in": None,
            "mature_width_max_in": None,
            "annual_growth_height_in": None,
            "annual_growth_width_in": None,
            "growth_rate_code": "unknown",
            "growth_habit": None,
            "root_behavior": None,
            "spread_aggressiveness": None,
            "pruning_response": None,
            "transplant_tolerance": None,
            "container_tolerance": None,
            "trellis_needed": None,
            "support_type": None,
        },
        "propagation_methods": [],
        "flowering": {
            "flowering_bool": None,
            "flower_color": None,
            "flower_size": None,
            "bloom_start_week": None,
            "bloom_end_week": None,
            "bloom_duration_days": None,
            "flower_abundance": None,
            "flower_fragrance_strength": None,
            "pollinator_value": None,
            "nectar_value": None,
            "pollen_value": None,
            "attracts_bees": None,
            "attracts_butterflies": None,
            "attracts_hummingbirds": None,
            "larval_host": None,
            "native_pollinator_value": None,
        },
        "fruiting": {
            "fruiting_bool": None,
            "fruit_color": None,
            "fruit_size": None,
            "fruit_flavor": None,
            "fruiting_start_age_years": None,
            "yield_lb_per_plant_year_min": None,
            "yield_lb_per_plant_year_max": None,
            "harvest_window_start_week": None,
            "harvest_window_end_week": None,
            "fruit_drop_behavior": None,
            "wildlife_attraction": None,
            "first_harvest_time_from_planting_days": None,
            "productive_years_min": None,
            "productive_years_max": None,
            "harvest_frequency": None,
            "preservation_uses": None,
            "edible_parts": [],
            "medicinal_parts": [],
            "fodder_parts": [],
        },
        "soil": {
            "drainage_requirement": None,
            "organic_matter_preference": None,
            "compaction_tolerance_code": "unknown",
            "rocky_soil_tolerance_code": "unknown",
            "ph_min": None,
            "ph_max": None,
            "ph_ideal_min": None,
            "ph_ideal_max": None,
            "ph_sensitivity_code": "unknown",
            "fertility_need": None,
            "nitrogen_need": None,
            "phosphorus_need": None,
            "potassium_need": None,
            "calcium_sensitivity_code": "unknown",
            "soil_oxygen_need": None,
            "mycorrhizal_association_notes": None,
            "mulch_preference": None,
            "mulch_depth_preference_in": None,
            "waterlogging_sensitivity_code": "unknown",
            "texture_preferences": {},
            "preferred_soil_texture_codes": [],
            "soil_texture_summary": None,
        },
        "water": {
            "water_need_level": "medium",
            "drought_tolerance_code": "unknown",
            "moisture_sensitivity_code": "unknown",
            "preferred_irrigation_method": None,
            "root_zone_depth_in": None,
            "container_water_multiplier": None,
            "mulched_water_reduction_factor": None,
            "summer_heat_adjustment_factor": None,
        },
        "water_establishment": [],
        "water_seasonal": [],
        "ecology": {
            "invasive_risk_code": "unknown",
            "wildlife_food_value": None,
            "erosion_control_value": None,
            "biomass_value": None,
            "compost_value": None,
            "chop_drop_value": None,
        },
        "maintenance": {
            "pruning_frequency": None,
            "deadheading_helpful": None,
            "division_interval_years": None,
            "staking_needed": None,
            "suckering_management": None,
            "cleanup_intensity": None,
            "disease_susceptibility_level": None,
            "pest_susceptibility_level": None,
            "humidity_disease_risk": None,
            "air_flow_importance": None,
        },
        "safety": [
            {
                "subject_type_code": "human",
                "safety_level_code": "unknown",
                "toxic_parts": [],
                "condition_notes": "Safety not yet curated; draft record only.",
                "symptoms": None,
                "evidence_source_type": "curation_needed",
                "safe_use_notes": "Do not use edible, medicinal, livestock, or pet-safety assumptions until source-enriched.",
            }
        ],
        "relationships": [],
        "phenology_templates": [],
        "zone_profiles": [],
        "care_events": [],
        "sources": [
            {
                "source_ref": source_ref,
                "source_name": "Garden.io March 2026 starter workbook",
                "source_type": "internal_curation",
                "publisher": "Garden.io",
                "author": None,
                "source_url": None,
                "citation_text": clean_text(row.get("source")) or "Garden.io starter workbook import row.",
                "published_on": None,
                "credibility_score": 0.3,
                "license": None,
                "notes": source_note(row),
                "last_reviewed_at": now,
            }
        ],
        "claims": [
            {
                "claim_type": "profile.workbook_presence",
                "value_json": {
                    "plant_name": clean_text(row.get("plant_name")),
                    "catalog_slug": slug,
                    "sheet_name": row.get("_sheet_name"),
                    "zone_name": clean_text(row.get("zone_name")) or clean_text(row.get("candidate_zone")),
                    "bed_name": clean_text(row.get("bed_name")) or clean_text(row.get("candidate_bed")),
                    "quantity": clean_text(row.get("quantity")),
                    "status": clean_text(row.get("status")),
                    "notes": clean_text(row.get("notes")),
                },
                "evidence_strength_code": "unknown",
                "confidence_score": 0.3,
                "source_ref": source_ref,
                "source_quote_or_excerpt": source_note(row),
                "source_url": None,
                "reviewed_by_human": False,
                "review_status": "needs_more_evidence",
                "region_scope": None,
                "cultivar_scope": None,
                "ai_generated_summary": False,
                "human_verified": False,
                "conflict_flag": False,
                "region_specific_conflict_notes": None,
            }
        ],
        "images": [
            {
                "source_ref": None,
                "stage_code": None,
                "image_url": "/art/specimen-herbarium-sheet.svg",
                "storage_key": "art/specimen-herbarium-sheet.svg",
                "mime_type": "image/svg+xml",
                "width_px": None,
                "height_px": None,
                "attribution_text": "Garden.io placeholder specimen illustration",
                "license": "internal placeholder",
                "is_primary": True,
                "is_public": True,
            }
        ],
        "ratings": [],
        "import_notes": "Draft starter-batch record generated deterministically from workbook provenance only. Requires source-enrichment before publication.",
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate draft catalog plant records from the starter workbook.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--limit", type=int, default=5)
    parser.add_argument("--prefer-explicit-slugs", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    rows = workbook_candidates(args.workbook, args.prefer_explicit_slugs)[: args.limit]
    if not rows:
        raise SystemExit("No workbook rows matched the selected criteria.")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for row in rows:
        record = draft_record(row)
        path = args.output_dir / f"{record['slug']}.json"
        if path.exists() and not args.overwrite:
            continue
        path.write_text(json.dumps(record, indent=2, sort_keys=False) + "\n", encoding="utf-8")
        written.append(path)

    print(f"Wrote {len(written)} draft record(s) to {args.output_dir}")
    for path in written:
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

