#!/usr/bin/env python3
"""Import the Garden.io starter workbook into the private-beta Supabase schema.

The script is intentionally idempotent. It applies the private-beta schema,
creates/fetches the target auth user, creates catalogue records for imported
plants that are not in the seed catalogue, then upserts the property hierarchy
and attached records from the starter workbook.
"""

from __future__ import annotations

import argparse
import os
import re
import subprocess
import tempfile
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path("/Users/preston/Downloads/garden_io_starter_import_march_2026.xlsx")
DEFAULT_EMAIL = "prestonpope192@gmail.com"
DEFAULT_FULL_NAME = "Preston Pope"
NAMESPACE = uuid.UUID("d7c33a95-7776-48c3-8846-03c3f63002aa")


def stable_uuid(*parts: Any) -> str:
    return str(uuid.uuid5(NAMESPACE, "::".join(str(part) for part in parts)))


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "imported-plant"


def sql(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, datetime):
        return "'" + value.isoformat(sep=" ") + "'"
    if isinstance(value, date):
        return "'" + value.isoformat() + "'"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value)
    if not text or text.startswith("[TO FILL"):
        return "null"
    return "'" + text.replace("'", "''") + "'"


def read_sheet(workbook: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(workbook, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell).strip() for cell in rows[0] if cell is not None]
    records: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        values = raw_row[: len(headers)]
        if not any(value not in (None, "") for value in values):
            continue
        records.append(dict(zip(headers, values)))
    return records


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.startswith("[TO FILL"):
        return None
    return text


def join_notes(*parts: Any) -> str | None:
    clean_parts = [clean_text(part) for part in parts]
    clean_parts = [part for part in clean_parts if part]
    return "\n".join(clean_parts) if clean_parts else None


def parse_quantity(value: Any) -> str:
    if isinstance(value, (int, float)) and value > 0:
        return str(value)

    text = clean_text(value)
    if not text:
        return "1"

    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return "1"

    quantity = float(match.group(0))
    return match.group(0) if quantity > 0 else "1"


def build_import_sql(workbook_path: Path, email: str, full_name: str) -> str:
    account = {row["field"]: row for row in read_sheet(workbook_path, "account_basics")}
    property_rows = read_sheet(workbook_path, "property")
    zones = read_sheet(workbook_path, "zones")
    beds = read_sheet(workbook_path, "beds")
    plants = read_sheet(workbook_path, "plants")
    observations = read_sheet(workbook_path, "observations")
    tasks = read_sheet(workbook_path, "tasks")
    wishlist = read_sheet(workbook_path, "wishlist")

    property_row = property_rows[0]
    user_id = stable_uuid("auth-user", email.lower())
    property_id = stable_uuid("property", email.lower(), "starter")
    property_name = clean_text(property_row.get("name")) or "Preston Homestead"
    property_label = clean_text(property_row.get("label")) or "homestead"
    user_ref = "(select id from garden_import_target_user)"
    property_notes = join_notes(
        property_row.get("notes"),
        "Imported from garden_io_starter_import_march_2026.xlsx.",
        "Starter workbook source: Plants list as of March 2026."
    )

    zone_ids = {clean_text(row.get("zone_name")): stable_uuid("zone", clean_text(row.get("zone_name"))) for row in zones}
    bed_ids = {
        (clean_text(row.get("zone_name")), clean_text(row.get("bed_name"))): stable_uuid(
            "bed", clean_text(row.get("zone_name")), clean_text(row.get("bed_name"))
        )
        for row in beds
    }

    def catalog_slug_for(row: dict[str, Any], index: int, prefix: str) -> str:
        explicit = clean_text(row.get("catalog_slug"))
        if explicit:
            return slugify(explicit)
        return slugify(f"{row.get('plant_name')}") or f"{prefix}-{index + 1}"

    catalog_rows: dict[str, str] = {}
    plant_ids: dict[tuple[str, int], str] = {}
    first_plant_by_name: dict[str, str] = {}
    for index, row in enumerate(plants):
        plant_name = clean_text(row.get("plant_name"))
        if not plant_name:
            continue
        slug = catalog_slug_for(row, index, "plant")
        catalog_rows.setdefault(slug, plant_name)
        plant_id = stable_uuid("plant", index, plant_name, row.get("zone_name"), row.get("bed_name"))
        plant_ids[(plant_name, index)] = plant_id
        first_plant_by_name.setdefault(plant_name.lower(), plant_id)

    for index, row in enumerate(wishlist):
        plant_name = clean_text(row.get("plant_name"))
        if not plant_name:
            continue
        slug = catalog_slug_for(row, index, "wishlist")
        catalog_rows.setdefault(slug, plant_name)

    catalog_taxon_ids = {slug: stable_uuid("catalog-taxon", slug) for slug in catalog_rows}
    catalog_profile_ids = {slug: stable_uuid("catalog-profile", slug) for slug in catalog_rows}

    lines = [
        "begin;",
        "",
        "create extension if not exists pgcrypto;",
        "create temp table garden_import_target_user (id uuid primary key);",
        "",
        "do $$",
        "declare",
        f"  target_user_id uuid := {sql(user_id)}::uuid;",
        "begin",
        f"  select id into target_user_id from auth.users where lower(email) = lower({sql(email)}) limit 1;",
        "  if target_user_id is null then",
        f"    target_user_id := {sql(user_id)}::uuid;",
        "    insert into auth.users (",
        "      id, instance_id, aud, role, email, encrypted_password, email_confirmed_at,",
        "      confirmation_sent_at, raw_app_meta_data, raw_user_meta_data, created_at, updated_at",
        "    ) values (",
        f"      target_user_id, '00000000-0000-0000-0000-000000000000', 'authenticated',",
        f"      'authenticated', {sql(email)}, crypt(gen_random_uuid()::text, gen_salt('bf')), now(),",
        f"      now(), '{{\"provider\":\"email\",\"providers\":[\"email\"]}}'::jsonb,",
        f"      jsonb_build_object('name', {sql(full_name)}), now(), now()",
        "    );",
        "  end if;",
        "",
        "  insert into auth.identities (id, user_id, provider_id, identity_data, provider, last_sign_in_at, created_at, updated_at)",
        "  values (",
        "    target_user_id, target_user_id, " + sql(email) + ",",
        "    jsonb_build_object('sub', target_user_id::text, 'email', " + sql(email) + "),",
        "    'email', now(), now(), now()",
        "  ) on conflict do nothing;",
        "",
        "  insert into garden_import_target_user (id) values (target_user_id);",
        "end $$;",
        "",
        "insert into public.garden_profiles (id, email, full_name)",
        f"values ({user_ref}, {sql(email)}, {sql(full_name)})",
        "on conflict (id) do update set",
        "  email = excluded.email,",
        "  full_name = excluded.full_name,",
        "  updated_at = now();",
        "",
    ]

    def catalog_profile_id_ref(slug: str) -> str:
        return f"(select id from catalog.plant_profiles where slug = {sql(slug)})"

    for slug, plant_name in sorted(catalog_rows.items()):
        taxon_id = catalog_taxon_ids[slug]
        profile_id = catalog_profile_ids[slug]
        lines.extend(
            [
                "insert into catalog.plant_taxa (id, genus_name, botanical_name_full, taxon_rank, origin_type)",
                "values (",
                f"  {sql(taxon_id)}::uuid, {sql(slug.replace('-', '_'))}, {sql(plant_name)}, 'unknown', 'unknown'",
                ")",
                "on conflict (id) do update set",
                "  botanical_name_full = excluded.botanical_name_full,",
                "  updated_at = now();",
                "",
                "insert into catalog.plant_names (plant_taxon_id, name, name_type, locale, is_primary)",
                "values (",
                f"  {sql(taxon_id)}::uuid, {sql(plant_name)}, 'common', 'en', true",
                ")",
                "on conflict (plant_taxon_id, (lower(name)), name_type, locale) do update set",
                "  is_primary = excluded.is_primary,",
                "  updated_at = now();",
                "",
                "insert into catalog.plant_profiles (",
                "  id, plant_taxon_id, slug, display_name, plant_type_code, lifecycle_type,",
                "  generation_status, is_published, review_status, evidence_count, source_count, human_verified",
                ") values (",
                f"  {sql(profile_id)}::uuid, {sql(taxon_id)}::uuid, {sql(slug)}, {sql(plant_name)},",
                "  'forb', 'unknown', 'community_generated', false, 'draft', 0, 0, false",
                ")",
                "on conflict (id) do update set",
                "  slug = excluded.slug,",
                "  display_name = excluded.display_name,",
                "  updated_at = now();",
                "",
                "insert into catalog.plant_profile_narratives (plant_profile_id, locale, short_description)",
                "values (",
                f"  {sql(profile_id)}::uuid, 'en', {sql(f'{plant_name} imported from the March 2026 Garden.io starter workbook.')}",
                ")",
                "on conflict (plant_profile_id, locale) do update set",
                "  short_description = excluded.short_description,",
                "  updated_at = now();",
                "",
            ]
        )

    lines.extend(
        [
            "insert into public.garden_properties (id, owner_user_id, name, label, region, growing_zone, season, notes)",
            "values (",
            f"  {sql(property_id)}::uuid, {user_ref}, {sql(property_name)}, {sql(property_label)},",
            f"  {sql(property_row.get('region'))}, {sql(property_row.get('growing_zone'))},",
            f"  {sql(property_row.get('season'))}, {sql(property_notes)}",
            ")",
            "on conflict (id) do update set",
            "  name = excluded.name,",
            "  label = excluded.label,",
            "  region = excluded.region,",
            "  growing_zone = excluded.growing_zone,",
            "  season = excluded.season,",
            "  notes = excluded.notes,",
            "  updated_at = now();",
            "",
        ]
    )

    for index, row in enumerate(zones):
        zone_name = clean_text(row.get("zone_name"))
        if not zone_name:
            continue
        lines.extend(
            [
                "insert into public.garden_zones (id, property_id, name, purpose, light, water, notes, sort_order)",
                "values (",
                f"  {sql(zone_ids[zone_name])}::uuid, {sql(property_id)}::uuid, {sql(zone_name)},",
                f"  {sql(row.get('purpose'))}, {sql(row.get('light'))}, {sql(row.get('water'))},",
                f"  {sql(join_notes(row.get('notes'), row.get('source')))}, {index}",
                ")",
                "on conflict (id) do update set",
                "  name = excluded.name, purpose = excluded.purpose, light = excluded.light,",
                "  water = excluded.water, notes = excluded.notes, sort_order = excluded.sort_order, updated_at = now();",
                "",
            ]
        )

    for index, row in enumerate(beds):
        zone_name = clean_text(row.get("zone_name"))
        bed_name = clean_text(row.get("bed_name"))
        if not zone_name or not bed_name:
            continue
        lines.extend(
            [
                "insert into public.garden_beds (id, property_id, zone_id, name, sun, water, soil, notes, sort_order)",
                "values (",
                f"  {sql(bed_ids[(zone_name, bed_name)])}::uuid, {sql(property_id)}::uuid,",
                f"  {sql(zone_ids[zone_name])}::uuid, {sql(bed_name)}, {sql(row.get('sun'))},",
                f"  {sql(row.get('water'))}, {sql(row.get('soil'))},",
                f"  {sql(join_notes(row.get('bed_type'), row.get('notes'), row.get('source')))}, {index}",
                ")",
                "on conflict (id) do update set",
                "  name = excluded.name, sun = excluded.sun, water = excluded.water, soil = excluded.soil,",
                "  notes = excluded.notes, sort_order = excluded.sort_order, updated_at = now();",
                "",
            ]
        )

    for index, row in enumerate(plants):
        plant_name = clean_text(row.get("plant_name"))
        zone_name = clean_text(row.get("zone_name"))
        bed_name = clean_text(row.get("bed_name"))
        if not plant_name or not zone_name or not bed_name:
            continue
        slug = catalog_slug_for(row, index, "plant")
        plant_id = plant_ids[(plant_name, index)]
        quantity = parse_quantity(row.get("quantity"))
        plant_notes = join_notes(
            row.get("notes"),
            f"Source type: {clean_text(row.get('source_type'))}" if clean_text(row.get("source_type")) else None,
            row.get("source"),
        )
        lines.extend(
            [
                "insert into public.garden_plant_instances (id, property_id, zone_id, bed_id, plant_profile_id, quantity, status, planted_on, notes)",
                "values (",
                f"  {sql(plant_id)}::uuid, {sql(property_id)}::uuid, {sql(zone_ids[zone_name])}::uuid,",
                f"  {sql(bed_ids[(zone_name, bed_name)])}::uuid, {catalog_profile_id_ref(slug)},",
                f"  {quantity}, {sql(clean_text(row.get('status')) or 'growing')}, {sql(row.get('planted_on'))}, {sql(plant_notes)}",
                ")",
                "on conflict (id) do update set",
                "  plant_profile_id = excluded.plant_profile_id, quantity = excluded.quantity, status = excluded.status,",
                "  planted_on = excluded.planted_on, notes = excluded.notes, updated_at = now();",
                "",
            ]
        )

    for index, row in enumerate(observations):
        note = clean_text(row.get("note"))
        if not note:
            continue
        zone_name = clean_text(row.get("zone_name"))
        bed_name = clean_text(row.get("bed_name"))
        plant_name = clean_text(row.get("plant_name"))
        observation_id = stable_uuid("observation", index, note)
        lines.extend(
            [
                "insert into public.garden_observations (id, property_id, zone_id, bed_id, plant_instance_id, note, observed_at)",
                "values (",
                f"  {sql(observation_id)}::uuid, {sql(property_id)}::uuid,",
                f"  {sql(zone_ids.get(zone_name)) + '::uuid' if zone_name in zone_ids else 'null'},",
                f"  {sql(bed_ids.get((zone_name, bed_name))) + '::uuid' if (zone_name, bed_name) in bed_ids else 'null'},",
                f"  {sql(first_plant_by_name.get(plant_name.lower())) + '::uuid' if plant_name and plant_name.lower() in first_plant_by_name else 'null'},",
                f"  {sql(join_notes(note, row.get('observation_type'), row.get('source')))}, {sql(row.get('observed_at'))}",
                ")",
                "on conflict (id) do update set",
                "  zone_id = excluded.zone_id, bed_id = excluded.bed_id, plant_instance_id = excluded.plant_instance_id,",
                "  note = excluded.note, observed_at = excluded.observed_at, updated_at = now();",
                "",
            ]
        )

    for index, row in enumerate(tasks):
        title = clean_text(row.get("title"))
        if not title:
            continue
        zone_name = clean_text(row.get("zone_name"))
        bed_name = clean_text(row.get("bed_name"))
        plant_name = clean_text(row.get("plant_name"))
        task_id = stable_uuid("task", index, title)
        status = clean_text(row.get("status")) or "open"
        lines.extend(
            [
                "insert into public.garden_tasks (id, property_id, zone_id, bed_id, plant_instance_id, title, notes, due_on, status, completed_at)",
                "values (",
                f"  {sql(task_id)}::uuid, {sql(property_id)}::uuid,",
                f"  {sql(zone_ids.get(zone_name)) + '::uuid' if zone_name in zone_ids else 'null'},",
                f"  {sql(bed_ids.get((zone_name, bed_name))) + '::uuid' if (zone_name, bed_name) in bed_ids else 'null'},",
                f"  {sql(first_plant_by_name.get(plant_name.lower())) + '::uuid' if plant_name and plant_name.lower() in first_plant_by_name else 'null'},",
                f"  {sql(title)}, {sql(join_notes(row.get('notes'), row.get('priority'), row.get('source')))},",
                f"  {sql(row.get('due_on'))}, {sql(status)}, {'now()' if status == 'done' else 'null'}",
                ")",
                "on conflict (id) do update set",
                "  zone_id = excluded.zone_id, bed_id = excluded.bed_id, plant_instance_id = excluded.plant_instance_id,",
                "  title = excluded.title, notes = excluded.notes, due_on = excluded.due_on, status = excluded.status,",
                "  completed_at = excluded.completed_at, updated_at = now();",
                "",
            ]
        )

    for index, row in enumerate(wishlist):
        plant_name = clean_text(row.get("plant_name"))
        if not plant_name:
            continue
        slug = catalog_slug_for(row, index, "wishlist")
        wishlist_id = stable_uuid("wishlist", index, plant_name, row.get("candidate_zone"), row.get("candidate_bed"))
        notes = join_notes(
            f"Candidate zone: {clean_text(row.get('candidate_zone'))}" if clean_text(row.get("candidate_zone")) else None,
            f"Candidate bed: {clean_text(row.get('candidate_bed'))}" if clean_text(row.get("candidate_bed")) else None,
            row.get("notes"),
            row.get("source"),
        )
        lines.extend(
            [
                "insert into public.garden_wishlist (id, owner_user_id, plant_profile_id, notes)",
                "values (",
                f"  {sql(wishlist_id)}::uuid, {user_ref}, {catalog_profile_id_ref(slug)}, {sql(notes)}",
                ")",
                "on conflict (owner_user_id, plant_profile_id) do update set",
                "  notes = concat_ws(E'\\n\\n', public.garden_wishlist.notes, excluded.notes),",
                "  updated_at = now();",
                "",
            ]
        )

    lines.extend(
        [
            "commit;",
            "",
            "select 'garden_profiles' as table_name, count(*) from public.garden_profiles where id = "
            + user_ref,
            "union all select 'garden_properties', count(*) from public.garden_properties where owner_user_id = "
            + user_ref,
            "union all select 'garden_zones', count(*) from public.garden_zones where property_id = "
            + sql(property_id)
            + "::uuid",
            "union all select 'garden_beds', count(*) from public.garden_beds where property_id = "
            + sql(property_id)
            + "::uuid",
            "union all select 'garden_plant_instances', count(*) from public.garden_plant_instances where property_id = "
            + sql(property_id)
            + "::uuid",
            "union all select 'garden_observations', count(*) from public.garden_observations where property_id = "
            + sql(property_id)
            + "::uuid",
            "union all select 'garden_tasks', count(*) from public.garden_tasks where property_id = "
            + sql(property_id)
            + "::uuid",
            "union all select 'garden_wishlist', count(*) from public.garden_wishlist where owner_user_id = "
            + user_ref
            + ";",
            "",
        ]
    )
    return "\n".join(lines)


def run_psql(db_url: str, sql_path: Path) -> None:
    subprocess.run(["psql", db_url, "-v", "ON_ERROR_STOP=1", "-f", str(sql_path)], check=True)


def write_canonical_schema_without_rls(output_path: Path) -> None:
    schema_path = ROOT / "docs/product/specs/sql/10-garden-postgres-ddl.sql"
    schema_sql = schema_path.read_text()
    marker = "-- ============================================================================\n-- Row-level security policies"
    marker_index = schema_sql.find(marker)
    if marker_index == -1:
        marker_index = schema_sql.find("alter table core.properties enable row level security;")
    if marker_index == -1:
        raise SystemExit("Unable to find row-level security section in canonical schema")

    output_path.write_text(schema_sql[:marker_index].rstrip() + "\n\ncommit;\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--email", default=DEFAULT_EMAIL)
    parser.add_argument("--full-name", default=DEFAULT_FULL_NAME)
    parser.add_argument("--output-sql", type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    import_sql = build_import_sql(args.workbook, args.email, args.full_name)

    if args.output_sql:
        args.output_sql.parent.mkdir(parents=True, exist_ok=True)
        args.output_sql.write_text(import_sql)
        print(f"Wrote import SQL to {args.output_sql}")

    if not args.apply:
        return

    db_url = os.environ.get("SUPABASE_DB_URL")
    if not db_url:
        raise SystemExit("SUPABASE_DB_URL is required when using --apply")

    schema_path = ROOT / "supabase/sql/20-private-beta-mvp.sql"
    quantity_path = ROOT / "supabase/sql/23-private-beta-plant-instance-quantity-placement.sql"
    canonical_bridge_path = ROOT / "supabase/sql/26-private-beta-canonical-catalog-poc.sql"
    canonical_validation_path = ROOT / "supabase/sql/27-private-beta-canonical-catalog-validate.sql"
    rating_scales_path = ROOT / "supabase/sql/28-private-beta-normalized-plant-rating-scales.sql"
    profile_header_refactor_path = ROOT / "supabase/sql/29-private-beta-profile-header-refactor.sql"
    drop_mvp_catalog_path = ROOT / "supabase/sql/30-private-beta-drop-mvp-catalog-table.sql"
    merge_soil_texture_path = ROOT / "supabase/sql/31-private-beta-merge-soil-texture-preferences.sql"
    merge_propagation_path = ROOT / "supabase/sql/32-private-beta-merge-propagation-profile-methods.sql"
    water_quality_path = ROOT / "supabase/sql/33-private-beta-water-quality-tracking.sql"
    shallow_well_seed_path = ROOT / "supabase/sql/35-private-beta-shallow-well-water-quality-seed.sql"
    chicken_pond_seed_path = ROOT / "supabase/sql/36-private-beta-chicken-pond-water-quality-seed.sql"
    shallow_well_2025_seed_path = ROOT / "supabase/sql/37-private-beta-shallow-well-2025-water-quality-seed.sql"
    validation_path = ROOT / "supabase/sql/21-private-beta-mvp-validate.sql"
    water_quality_validation_path = ROOT / "supabase/sql/34-private-beta-water-quality-validate.sql"
    with tempfile.TemporaryDirectory() as tmp_dir:
        import_path = Path(tmp_dir) / "garden_starter_import.sql"
        canonical_schema_path = Path(tmp_dir) / "garden_canonical_schema_no_rls.sql"
        import_path.write_text(import_sql)
        write_canonical_schema_without_rls(canonical_schema_path)
        run_psql(db_url, schema_path)
        run_psql(db_url, canonical_schema_path)
        run_psql(db_url, profile_header_refactor_path)
        run_psql(db_url, drop_mvp_catalog_path)
        run_psql(db_url, merge_soil_texture_path)
        run_psql(db_url, merge_propagation_path)
        run_psql(db_url, water_quality_path)
        run_psql(db_url, import_path)
        run_psql(db_url, shallow_well_seed_path)
        run_psql(db_url, chicken_pond_seed_path)
        run_psql(db_url, shallow_well_2025_seed_path)
        run_psql(db_url, quantity_path)
        run_psql(db_url, canonical_bridge_path)
        run_psql(db_url, rating_scales_path)
        run_psql(db_url, validation_path)
        run_psql(db_url, water_quality_validation_path)
        run_psql(db_url, canonical_validation_path)


if __name__ == "__main__":
    main()
